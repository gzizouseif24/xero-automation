"""
File parser interfaces and implementations for Xero Payroll Automation system.

This module defines the abstract base class for file parsers and provides
concrete implementations for different Excel file formats used in the payroll process.
"""

from abc import ABC, abstractmethod
from typing import Dict, Any, List, Optional
from pathlib import Path
from datetime import date, datetime, timedelta
import openpyxl
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from config.settings import (
    HOUR_TYPE_TO_XERO_MAPPING, 
    HOLIDAY_HOURS_PER_DAY, 
    SUPPORTED_FILE_EXTENSIONS,
    SITE_TIMESHEET_CONFIG,
    TRAVEL_TIME_CONFIG,
    OVERTIME_RATES_CONFIG
)



class FileParser(ABC):
    """
    Abstract base class for file parsers.
    
    This class defines the interface that all file parsers must implement
    to ensure consistent behavior across different file formats.
    """
    
    @abstractmethod
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse the file and return structured data.
        
        Args:
            file_path: Path to the file to parse
            
        Returns:
            Dictionary containing parsed data in a standardized format
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid or data is malformed
            PermissionError: If the file cannot be read
        """
        pass
    
    @abstractmethod
    def validate_format(self, file_path: str) -> bool:
        """
        Validate that the file format is correct before parsing.
        
        Args:
            file_path: Path to the file to validate
            
        Returns:
            True if the file format is valid, False otherwise
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            PermissionError: If the file cannot be read
        """
        pass
    
    def _validate_file_exists(self, file_path: str) -> None:
        """
        Validate that the file exists and is readable.
        
        Args:
            file_path: Path to the file to validate
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            PermissionError: If the file cannot be read
        """
        path = Path(file_path)
        
        if not path.exists():
            raise FileNotFoundError(f"File not found: {file_path}")
        
        if not path.is_file():
            raise ValueError(f"Path is not a file: {file_path}")
        
        # Try to open the file to check permissions
        try:
            with open(file_path, 'rb') as f:
                f.read(1)  # Read just one byte to test access
        except PermissionError:
            raise PermissionError(f"Cannot read file: {file_path}")
    
    def _load_excel_workbook(self, file_path: str) -> Workbook:
        """
        Load an Excel workbook from the given file path.
        
        Args:
            file_path: Path to the Excel file
            
        Returns:
            Loaded openpyxl Workbook object
            
        Raises:
            ValueError: If the file is not a valid Excel file
        """
        try:
            return openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            raise ValueError(f"Invalid Excel file format: {file_path}. Error: {str(e)}")
    
    def _get_worksheet_by_name(self, workbook: Workbook, sheet_name: str) -> Optional[Worksheet]:
        """
        Get a worksheet by name from the workbook.
        
        Args:
            workbook: The Excel workbook
            sheet_name: Name of the worksheet to retrieve
            
        Returns:
            Worksheet object if found, None otherwise
        """
        return workbook.get(sheet_name)
    
    def _get_first_worksheet(self, workbook: Workbook) -> Worksheet:
        """
        Get the first worksheet from the workbook.
        
        Args:
            workbook: The Excel workbook
            
        Returns:
            First worksheet in the workbook
            
        Raises:
            ValueError: If the workbook has no worksheets
        """
        if not workbook.worksheets:
            raise ValueError("Workbook contains no worksheets")
        
        return workbook.worksheets[0]
    
    def _cell_value_to_string(self, cell_value: Any) -> str:
        """
        Convert a cell value to a string, handling None and various data types.
        
        Args:
            cell_value: The cell value from openpyxl
            
        Returns:
            String representation of the cell value
        """
        if cell_value is None:
            return ""
        
        return str(cell_value).strip()
    
    def _cell_value_to_float(self, cell_value: Any, default: float = 0.0) -> float:
        """
        Convert a cell value to a float, handling various data types and formats.
        
        Args:
            cell_value: The cell value from openpyxl
            default: Default value to return if conversion fails
            
        Returns:
            Float representation of the cell value
        """
        if cell_value is None:
            return default
        
        # Handle numeric values directly
        if isinstance(cell_value, (int, float)):
            return float(cell_value)
        
        # Handle string values
        if isinstance(cell_value, str):
            cell_value = cell_value.strip()
            
            # Handle empty strings
            if not cell_value:
                return default
            
            # Handle comma as decimal separator
            cell_value = cell_value.replace(',', '.')
            
            # Try to convert to float
            try:
                return float(cell_value)
            except ValueError:
                return default
        
        return default
    
    def _find_header_row(self, worksheet: Worksheet, expected_headers: List[str], 
                        max_rows: int = 10) -> Optional[int]:
        """
        Find the row containing the expected headers.
        
        Args:
            worksheet: The Excel worksheet
            expected_headers: List of expected header names (case-insensitive)
            max_rows: Maximum number of rows to search
            
        Returns:
            Row number (1-based) containing the headers, or None if not found
        """
        expected_lower = [header.lower() for header in expected_headers]
        
        for row_num in range(1, min(max_rows + 1, worksheet.max_row + 1)):
            row_values = []
            for col_num in range(1, worksheet.max_column + 1):
                cell_value = worksheet.cell(row=row_num, column=col_num).value
                row_values.append(self._cell_value_to_string(cell_value).lower())
            
            # Check if all expected headers are present in this row
            headers_found = 0
            for expected_header in expected_lower:
                if expected_header in row_values:
                    headers_found += 1
            
            # If we found most of the expected headers, consider this the header row
            if headers_found >= len(expected_headers) * 0.8:  # 80% match threshold
                return row_num
        
        return None
    
    def _get_column_mapping(self, worksheet: Worksheet, header_row: int, 
                          expected_headers: List[str]) -> Dict[str, int]:
        """
        Create a mapping from header names to column numbers.
        
        Args:
            worksheet: The Excel worksheet
            header_row: Row number containing the headers
            expected_headers: List of expected header names
            
        Returns:
            Dictionary mapping header names to column numbers (1-based)
        """
        column_mapping = {}
        expected_lower = {header.lower(): header for header in expected_headers}
        
        for col_num in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=header_row, column=col_num).value
            header_name = self._cell_value_to_string(cell_value).lower()
            
            if header_name in expected_lower:
                original_header = expected_lower[header_name]
                column_mapping[original_header] = col_num
        
        return column_mapping


class ParseResult:
    """
    Standard result format returned by all file parsers.
    
    This class provides a consistent structure for parser results,
    making it easier to handle data from different file formats.
    """
    
    def __init__(self, success: bool, data: Optional[Dict[str, Any]] = None, 
                 errors: Optional[List[str]] = None):
        """
        Initialize a parse result.
        
        Args:
            success: Whether the parsing was successful
            data: Parsed data (if successful)
            errors: List of error messages (if unsuccessful)
        """
        self.success = success
        self.data = data or {}
        self.errors = errors or []
    
    def add_error(self, error_message: str) -> None:
        """Add an error message to the result."""
        self.errors.append(error_message)
        self.success = False
    
    def to_dict(self) -> Dict[str, Any]:
        """Convert the result to a dictionary."""
        return {
            "success": self.success,
            "data": self.data,
            "errors": self.errors
        }


# Standard data format returned by parsers
STANDARD_PARSER_FORMAT = {
    "file_type": "string",  # Type of file parsed (e.g., "site_timesheet", "travel_time", "overtime_rates")
    "pay_period_end_date": "date",  # End date of the pay period
    "employees": [
        {
            "employee_name": "string",  # Name of the employee as it appears in the file
            "entries": [
                {
                    "entry_date": "date",  # Date of the timesheet entry
                    "region_name": "string",  # Region/location name
                    "hours": "float",  # Number of hours worked
                    "hour_type": "string",  # Type of hours (REGULAR, OVERTIME, TRAVEL, HOLIDAY)
                    "overtime_rate": "float or null",  # Custom overtime rate if applicable
                    "notes": "string or null"  # Additional notes or comments
                }
            ]
        }
    ],
    "metadata": {
        "source_file": "string",  # Original file path
        "parsed_at": "datetime",  # When the file was parsed
        "total_employees": "int",  # Number of employees found
        "total_entries": "int",  # Total number of timesheet entries
        "date_range": {
            "start_date": "date",  # Earliest entry date
            "end_date": "date"  # Latest entry date
        }
    }
}


class SiteTimesheetParser(FileParser):
    """
    Parser for Site Timesheet Excel files.
    
    This parser handles the site timesheet format which contains employee names,
    dates, hours worked, and region information. It handles 'HOL' entries as
    8-hour holiday days, region overrides in hour cells, and overtime from totals columns.
    """
    
    def validate_format(self, file_path: str) -> bool:
        """
        Validate that the file is a valid site timesheet format.
        
        Args:
            file_path: Path to the file to validate
            
        Returns:
            True if the file format is valid, False otherwise
        """
        try:
            self._validate_file_exists(file_path)
            
            # Check if it's an Excel file
            if not any(file_path.lower().endswith(ext) for ext in SUPPORTED_FILE_EXTENSIONS):
                return False
            
            workbook = self._load_excel_workbook(file_path)
            
            # Check if any sheet looks like a site timesheet
            valid_sheet_found = False
            for worksheet in workbook.worksheets:
                if self._is_valid_site_timesheet_sheet(worksheet):
                    valid_sheet_found = True
                    break
            
            workbook.close()
            return valid_sheet_found
            
        except Exception:
            return False
    
    def _is_valid_site_timesheet_sheet(self, worksheet: Worksheet) -> bool:
        """
        Validate that a specific sheet is a valid site timesheet format.
        
        Args:
            worksheet: The Excel worksheet to validate
            
        Returns:
            True if the sheet format is valid, False otherwise
        """
        try:
            # Look for key indicators of a site timesheet:
            # - "WEEK ENDING:" or "REGION:" text
            # - Date pattern in row 9
            # - Employee names starting around row 11
            
            indicators_found = 0
            
            # Check for "WEEK ENDING:" and "REGION:" in first 10 rows
            for row in range(1, min(11, worksheet.max_row + 1)):
                for col in range(1, min(10, worksheet.max_column + 1)):
                    cell_value = self._cell_value_to_string(worksheet.cell(row=row, column=col).value).lower()
                    if 'week ending' in cell_value or 'region' in cell_value:
                        indicators_found += 1
                        break
            
            # Check for date pattern in row 9 (typical site timesheet structure)
            if worksheet.max_row >= 9:
                dates_found = 0
                for col in range(3, min(10, worksheet.max_column + 1)):  # Start from column C
                    cell_value = worksheet.cell(row=9, column=col).value
                    if self._parse_date_cell(cell_value):
                        dates_found += 1
                
                if dates_found >= 3:  # At least 3 dates suggests timesheet format
                    indicators_found += 1
            
            # Check for employee-like names in configured range
            employee_start_row = SITE_TIMESHEET_CONFIG["employee_start_row"]
            if worksheet.max_row >= employee_start_row:
                employee_names_found = 0
                for row in range(employee_start_row, min(employee_start_row + 5, worksheet.max_row + 1)):
                    cell_value = self._cell_value_to_string(worksheet.cell(row=row, column=2).value)
                    if cell_value and self._is_valid_employee_name(cell_value):
                        employee_names_found += 1
                
                if employee_names_found >= 1:  # At least 1 valid employee name
                    indicators_found += 1
            
            # Need at least 2 indicators to consider it a valid site timesheet
            return indicators_found >= 2
            
        except Exception:
            return False
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse the site timesheet file and return structured data.
        
        Args:
            file_path: Path to the site timesheet file
            
        Returns:
            Dictionary containing parsed timesheet data in standard format
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid or data is malformed
        """
        self._validate_file_exists(file_path)
        
        try:
            # 1. Initialize empty containers for aggregated results
            master_employees_data = {}
            all_dates = []
            total_entries = 0
            pay_period_end_date = None
            
            # 2. Load the workbook from file_path
            workbook = self._load_excel_workbook(file_path)
            
            # 3. Loop through every worksheet in workbook.worksheets
            for worksheet in workbook.worksheets:
                # 4. Call internal validation method for each worksheet
                if not self._is_valid_site_timesheet_sheet(worksheet):
                    # 5. If worksheet is not valid, silently continue to next sheet
                    continue
                
                # 6. If worksheet is valid, proceed to parse all data from that sheet
                try:
                    region_name = worksheet.title  # Use sheet name as default region
                    
                    # Extract pay period end date from the worksheet
                    if pay_period_end_date is None:
                        pay_period_end_date = self._extract_pay_period_date(worksheet)
                    
                    # Find the date row dynamically (more tolerant to layout variations)
                    date_row = self._find_date_row_in_sheet(worksheet, max_search_row=SITE_TIMESHEET_CONFIG["date_row_search_range"]) or 9
                    dates = self._extract_dates_from_row(worksheet, date_row)

                    if not dates:
                        # Try nearby rows as fallback
                        found = False
                        for try_row in range(max(1, date_row - 2), min(worksheet.max_row, date_row + 2) + 1):
                            dates_try = self._extract_dates_from_row(worksheet, try_row)
                            if dates_try:
                                dates = dates_try
                                found = True
                                break
                        if not found:
                            continue  # Skip sheets without valid dates
                    
                    all_dates.extend(dates)
                    
                    # Parse employee data starting from configured row
                    for row_num in range(SITE_TIMESHEET_CONFIG["employee_start_row"], worksheet.max_row + 1):
                        employee_name = self._cell_value_to_string(worksheet.cell(row=row_num, column=2).value)
                        
                        if not employee_name or not self._is_valid_employee_name(employee_name):
                            continue  # Skip empty rows or invalid employee names
                        
                        # 7. Merge data into master result containers
                        # Initialize employee data if not exists
                        if employee_name not in master_employees_data:
                            master_employees_data[employee_name] = []
                        
                        # Parse hours for each date (starting from column 3)
                        for col_idx, entry_date in enumerate(dates, start=3):
                            if col_idx > worksheet.max_column:
                                break
                            
                            hours_cell = worksheet.cell(row=row_num, column=col_idx)
                            hours_value = hours_cell.value
                            
                            if hours_value is None:
                                continue
                            
                            # Rule 1: Handle Region Overrides in Hour Cells
                            if isinstance(hours_value, str):
                                hours_value_upper = hours_value.upper().strip()
                                if hours_value_upper == 'HOL':
                                    # Handle 'HOL' entries as 8-hour holiday days
                                    hours = HOLIDAY_HOURS_PER_DAY
                                    hour_type = 'HOLIDAY'  # Use enum key, not display value
                                    current_region = region_name
                                else:
                                    # Any other string is a region override with configured regular hours
                                    hours = SITE_TIMESHEET_CONFIG["region_override_hours"]
                                    hour_type = 'REGULAR'  # Use enum key, not display value
                                    current_region = hours_value.strip()  # Use the string as region name
                            else:
                                # Numeric value - parse as regular hours
                                hours = self._cell_value_to_float(hours_value)
                                if hours <= 0:
                                    continue
                                hour_type = 'REGULAR'  # Use enum key, not display value
                                current_region = region_name
                            
                            # Add the entry
                            master_employees_data[employee_name].append({
                                "entry_date": entry_date,
                                "region_name": current_region,
                                "hours": hours,
                                "hour_type": hour_type,
                                "overtime_rate": None,
                                "notes": None
                            })
                            
                            total_entries += 1
                        
                        # Rule 2: Parse Overtime from the Totals Columns
                        # The overtime column is configurable (typically column L)
                        overtime_col = SITE_TIMESHEET_CONFIG["overtime_column"]
                        if overtime_col <= worksheet.max_column:
                            overtime_cell = worksheet.cell(row=row_num, column=overtime_col)
                            overtime_hours = self._cell_value_to_float(overtime_cell.value)
                            
                            if overtime_hours > 0:
                                # Create a separate overtime entry using pay period end date
                                overtime_date = pay_period_end_date if pay_period_end_date else (dates[-1] if dates else date.today())
                                master_employees_data[employee_name].append({
                                    "entry_date": overtime_date,
                                    "region_name": region_name,
                                    "hours": overtime_hours,
                                    "hour_type": 'OVERTIME',  # Use enum key, not display value
                                    "overtime_rate": None,
                                    "notes": "Overtime from totals column"
                                })
                                
                                total_entries += 1
                
                except Exception:
                    # If parsing this sheet fails, silently continue to next sheet
                    continue
            
            workbook.close()
            
            # 8. Build and return final, aggregated data dictionary
            # Convert to standard format
            employees_list = []
            for employee_name, entries in master_employees_data.items():
                employees_list.append({
                    "employee_name": employee_name,
                    "entries": entries
                })
            
            # Use extracted pay period end date or fallback to max date
            if pay_period_end_date is None:
                pay_period_end_date = max(all_dates) if all_dates else date.today()
            
            return {
                "file_type": "site_timesheet",
                "pay_period_end_date": pay_period_end_date,
                "employees": employees_list,
                "metadata": {
                    "source_file": file_path,
                    "parsed_at": datetime.now(),
                    "total_employees": len(employees_list),
                    "total_entries": total_entries,
                    "date_range": {
                        "start_date": min(all_dates) if all_dates else pay_period_end_date,
                        "end_date": pay_period_end_date
                    }
                }
            }
            
        except Exception as e:
            raise ValueError(f"Failed to parse site timesheet file {file_path}: {str(e)}")
    
    def _extract_pay_period_date(self, worksheet: Worksheet) -> Optional[date]:
        """
        Extract the pay period end date from a site timesheet worksheet.
        
        Args:
            worksheet: The Excel worksheet
            
        Returns:
            Pay period end date, or None if not found
        """
        # Look for "WEEK ENDING:" in the first few rows
        for row in range(1, 10):
            for col in range(1, 10):
                cell_value = self._cell_value_to_string(worksheet.cell(row=row, column=col).value).lower()
                if 'week ending' in cell_value:
                    # Check the next cell for the date
                    date_cell = worksheet.cell(row=row, column=col + 1)
                    return self._parse_date_cell(date_cell.value)
        
        return None

    def _find_date_row_in_sheet(self, worksheet: Worksheet, max_search_row: int = None) -> Optional[int]:
        """
        Find a row that contains multiple parseable dates (used as the date header row).
        Returns the row number or None if not found.
        """
        if max_search_row is None:
            max_search_row = SITE_TIMESHEET_CONFIG["date_row_search_range"]
        
        for row in range(1, min(max_search_row, worksheet.max_row) + 1):
            dates_found = 0
            for col in range(1, min(worksheet.max_column, 50) + 1):
                cell_value = worksheet.cell(row=row, column=col).value
                if self._parse_date_cell(cell_value):
                    dates_found += 1
                if dates_found >= 3:
                    return row

        return None
    
    def _extract_dates_from_row(self, worksheet: Worksheet, row_num: int) -> List[date]:
        """
        Extract dates from a specific row in the site timesheet.
        
        Args:
            worksheet: The Excel worksheet
            row_num: Row number containing dates
            
        Returns:
            List of dates found in the row
        """
        dates = []
        
        # Start from column 3 and extract dates
        for col in range(3, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_num, column=col).value
            parsed_date = self._parse_date_cell(cell_value)
            
            if parsed_date:
                dates.append(parsed_date)
            elif len(dates) > 0:
                # Stop when we hit a non-date after finding dates
                break
        
        return dates
    
    def _parse_date_cell(self, cell_value: Any) -> Optional[date]:
        """
        Parse a date cell value handling various formats.
        
        Args:
            cell_value: The cell value from openpyxl
            
        Returns:
            Parsed date object, or None if parsing fails
        """
        if cell_value is None:
            return None
        
        # Handle datetime objects (Excel dates)
        if isinstance(cell_value, datetime):
            return cell_value.date()
        
        if isinstance(cell_value, date):
            return cell_value
        
        # Handle string dates
        if isinstance(cell_value, str):
            cell_value = cell_value.strip()
            if not cell_value:
                return None
            
            # Try common date formats
            date_formats = [
                '%Y-%m-%d',
                '%d/%m/%Y',
                '%m/%d/%Y',
                '%d-%m-%Y',
                '%m-%d-%Y',
                '%d.%m.%Y',
                '%m.%d.%Y'
            ]
            
            for date_format in date_formats:
                try:
                    return datetime.strptime(cell_value, date_format).date()
                except ValueError:
                    continue
        
        # Handle numeric dates (Excel serial dates)
        if isinstance(cell_value, (int, float)):
            try:
                # Excel epoch starts at 1900-01-01, but has a leap year bug
                # openpyxl should handle this, but we'll be safe
                excel_epoch = datetime(1899, 12, 30)  # Adjusted for Excel's leap year bug
                return (excel_epoch + timedelta(days=cell_value)).date()
            except (ValueError, OverflowError):
                return None
        
        return None
    
    def _is_valid_employee_name(self, employee_name: str) -> bool:
        """
        Validate that a string represents a valid employee name.
        
        This filters out instruction text, headers, and other non-employee entries
        that might appear in the employee name column.
        
        Args:
            employee_name: The potential employee name to validate
            
        Returns:
            True if this appears to be a valid employee name, False otherwise
        """
        if not employee_name or not employee_name.strip():
            return False
        
        employee_name = employee_name.strip()
        
        # Filter out obvious instruction text and headers
        invalid_patterns = [
            'contractor name',
            'if you worked through break',
            'to temporary worker',
            'to be signed by supervisor',
            'standard hours and overtime',
            'certified as being correct',
            'unpaid breaks must not be included',
            'employee name',
            'name of employee',
            'total hours',
            'hours worked',
            'signature',
            'date signed',
            'supervisor',
            'manager approval'
        ]
        
        employee_lower = employee_name.lower()
        
        # Check for invalid patterns
        for pattern in invalid_patterns:
            if pattern in employee_lower:
                return False
        
        # Filter out entries that are too long (likely instruction text)
        if len(employee_name) > 50:
            return False
        
        # Filter out entries that contain too many common instruction words
        instruction_words = ['add', 'here', 'must', 'not', 'include', 'correct', 'payable', 'such', 'being']
        word_count = sum(1 for word in instruction_words if word in employee_lower)
        if word_count >= 3:
            return False
        
        # Filter out entries that are all uppercase and look like headers
        if employee_name.isupper() and len(employee_name.split()) <= 2:
            common_headers = ['CONTRACTOR NAME', 'EMPLOYEE NAME', 'TOTAL HOURS', 'SIGNATURE']
            if employee_name in common_headers:
                return False
        
        return True


class TravelTimeParser(FileParser):
    """
    Parser for Travel Time Excel files.
    
    This parser handles travel time data which contains travel hours that need
    to be distributed across working days for employees. Uses fixed column positions
    and handles comma as decimal separator.
    """
    
    def validate_format(self, file_path: str) -> bool:
        """
        Validate that the file is a valid travel time format.
        
        Args:
            file_path: Path to the file to validate
            
        Returns:
            True if the file format is valid, False otherwise
        """
        try:
            self._validate_file_exists(file_path)
            
            # Check if it's an Excel file
            if not any(file_path.lower().endswith(ext) for ext in SUPPORTED_FILE_EXTENSIONS):
                return False
            
            workbook = self._load_excel_workbook(file_path)
            
            # Check if any sheet looks like a travel time sheet
            valid_sheet_found = False
            for worksheet in workbook.worksheets:
                if self._is_valid_travel_time_sheet(worksheet):
                    valid_sheet_found = True
                    break
            
            workbook.close()
            return valid_sheet_found
            
        except Exception:
            return False
    
    def _is_valid_travel_time_sheet(self, worksheet: Worksheet) -> bool:
        """
        Validate that a specific sheet is a valid travel time format.
        
        Args:
            worksheet: The Excel worksheet to validate
            
        Returns:
            True if the sheet format is valid, False otherwise
        """
        try:
            # Look for key indicators of a travel time sheet:
            # - "Travel time" text in the sheet
            # - Valid employee names in column A
            # - Numeric hours in column D
            # - Avoid sheets with fake/test employee names
            
            indicators_found = 0
            
            # Check for "Travel time" indicator in the first few rows
            travel_text_found = False
            for row in range(1, min(10, worksheet.max_row + 1)):
                for col in range(1, min(5, worksheet.max_column + 1)):
                    cell_value = self._cell_value_to_string(worksheet.cell(row=row, column=col).value).lower()
                    if 'travel time' in cell_value:
                        travel_text_found = True
                        break
                if travel_text_found:
                    break
            
            if travel_text_found:
                indicators_found += 1
            
            # Check for realistic employee names in column A and hours in column D
            valid_employees = 0
            fake_employee_patterns = TRAVEL_TIME_CONFIG["fake_employee_patterns"]
            
            for row in range(TRAVEL_TIME_CONFIG["data_start_row"], min(20, worksheet.max_row + 1)):
                name_cell = self._cell_value_to_string(worksheet.cell(row=row, column=TRAVEL_TIME_CONFIG["employee_name_column"]).value)
                hours_cell = worksheet.cell(row=row, column=TRAVEL_TIME_CONFIG["hours_column"]).value
                
                if name_cell:
                    # Check if it looks like a fake name
                    is_fake = False
                    name_lower = name_cell.lower()
                    for pattern in fake_employee_patterns:
                        if pattern in name_lower:
                            is_fake = True
                            break
                    
                    # Also check for obvious test data patterns
                    if 'xxx' in name_lower or name_lower.startswith('test'):
                        is_fake = True
                    
                    if not is_fake and len(name_cell.split()) >= 2:  # At least first and last name
                        # Check if there's a corresponding hours value
                        hours_str = self._cell_value_to_string(hours_cell)
                        if hours_str:
                            hours_str = hours_str.replace(',', '.')  # Handle comma decimals
                            try:
                                hours_val = float(hours_str)
                                if hours_val > 0:
                                    valid_employees += 1
                            except (ValueError, TypeError):
                                pass
            
            if valid_employees >= 1:  # At least 1 valid employee with hours
                indicators_found += 1
            
            # Need at least 2 indicators for a valid travel time sheet
            return indicators_found >= 2
            
        except Exception:
            return False
    
    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse the travel time file and return structured data.
        
        Args:
            file_path: Path to the travel time file
            
        Returns:
            Dictionary containing parsed travel time data in standard format
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid or data is malformed
        """
        self._validate_file_exists(file_path)
        
        try:
            # 1. Initialize empty containers for aggregated results
            master_employees_data = {}
            all_dates = []
            total_entries = 0
            pay_period_end_date = None
            
            # 2. Load the workbook from file_path
            workbook = self._load_excel_workbook(file_path)
            
            # 3. Loop through every worksheet in workbook.worksheets
            for worksheet in workbook.worksheets:
                # 4. Call internal validation method for each worksheet
                if not self._is_valid_travel_time_sheet(worksheet):
                    # 5. If worksheet is not valid, silently continue to next sheet
                    continue
                
                # 6. If worksheet is valid, proceed to parse all data from that sheet
                try:
                    # Use configured column positions for travel time parsing
                    for row in range(TRAVEL_TIME_CONFIG["data_start_row"], worksheet.max_row + 1):
                        employee_name = self._cell_value_to_string(worksheet.cell(row=row, column=TRAVEL_TIME_CONFIG["employee_name_column"]).value)
                        site_name = self._cell_value_to_string(worksheet.cell(row=row, column=TRAVEL_TIME_CONFIG["site_name_column"]).value)
                        hours_cell = worksheet.cell(row=row, column=TRAVEL_TIME_CONFIG["hours_column"])
                        
                        if not employee_name:
                            continue
                        
                        # Skip fake/test employee names
                        fake_patterns = TRAVEL_TIME_CONFIG["fake_employee_patterns"]
                        is_fake = False
                        name_lower = employee_name.lower()
                        for pattern in fake_patterns:
                            if pattern in name_lower:
                                is_fake = True
                                break
                        
                        if is_fake:
                            continue
                        
                        # Handle comma as decimal separator for hours
                        hours_str = self._cell_value_to_string(hours_cell.value)
                        if not hours_str:
                            continue
                        
                        # Replace comma with period for decimal conversion
                        hours_str = hours_str.replace(',', '.')
                        try:
                            hours = float(hours_str)
                        except (ValueError, TypeError):
                            continue
                        
                        if hours <= 0:
                            continue
                        
                        # 7. Merge data into master result containers
                        if employee_name not in master_employees_data:
                            master_employees_data[employee_name] = []
                        
                        # Use placeholder date for travel time (will be adjusted by consolidator)
                        entry_date = date(1900, 1, 1)  # Placeholder date to be replaced by consolidator
                        if entry_date not in all_dates:
                            all_dates.append(entry_date)
                        
                        # Use site name as region, fallback to "Travel" if empty
                        region_name = site_name if site_name else "Travel"
                        
                        master_employees_data[employee_name].append({
                            "entry_date": entry_date,
                            "region_name": region_name,
                            "hours": hours,
                            "hour_type": "TRAVEL",  # Already correct - using enum key
                            "overtime_rate": None,
                            "notes": "Travel time entry"
                        })
                        
                        total_entries += 1
                        
                except Exception:
                    # If parsing this sheet fails, silently continue to next sheet
                    continue
            
            workbook.close()
            
            # 8. Build and return final, aggregated data dictionary
            employees_list = []
            for employee_name, entries in master_employees_data.items():
                employees_list.append({
                    "employee_name": employee_name,
                    "entries": entries
                })
            
            # Use placeholder date for travel time pay period (will be adjusted by consolidator)
            pay_period_end_date = date(1900, 1, 1)
            
            return {
                "file_type": "travel_time",
                "pay_period_end_date": pay_period_end_date,
                "employees": employees_list,
                "metadata": {
                    "source_file": file_path,
                    "parsed_at": datetime.now(),
                    "total_employees": len(employees_list),
                    "total_entries": total_entries,
                    "date_range": {
                        "start_date": min(all_dates) if all_dates else pay_period_end_date,
                        "end_date": max(all_dates) if all_dates else pay_period_end_date
                    }
                }
            }
            
        except Exception as e:
            raise ValueError(f"Failed to parse travel time file {file_path}: {str(e)}")


class OvertimeRatesParser(FileParser):
    """
    Parser for Employee Overtime Rates Excel files.
    
    This parser handles the master overtime rates file which contains employee names,
    overtime flags, and custom overtime rates for payroll calculations.
    It is robust against duplicate entries and handles mixed data types.
    """
    
    def validate_format(self, file_path: str) -> bool:
        """
        Validate that the file contains at least one valid overtime rates sheet.
        
        Args:
            file_path: Path to the file to validate
            
        Returns:
            True if the file format is valid, False otherwise
        """
        try:
            self._validate_file_exists(file_path)
            if not any(file_path.lower().endswith(ext) for ext in SUPPORTED_FILE_EXTENSIONS):
                return False
            
            workbook = self._load_excel_workbook(file_path)
            valid_sheet_found = False
            for worksheet in workbook.worksheets:
                if self._is_valid_overtime_sheet(worksheet):
                    valid_sheet_found = True
                    break
            
            workbook.close()
            return valid_sheet_found
            
        except Exception:
            return False

    def _is_valid_overtime_sheet(self, worksheet: Worksheet) -> bool:
        """
        Validate that a specific sheet is a valid overtime rates format.
        
        Args:
            worksheet: The Excel worksheet to validate
            
        Returns:
            True if the sheet format is valid, False otherwise
        """
        try:
            # Primary detection: look for the full title phrase in the first row
            first_row_values = []
            for col in range(1, min(worksheet.max_column, OVERTIME_RATES_CONFIG["max_search_columns"]) + 1):
                first_row_values.append(self._cell_value_to_string(worksheet.cell(row=1, column=col).value).lower())
            first_row_joined = ' '.join(first_row_values)
            if OVERTIME_RATES_CONFIG["exact_title_phrase"] in first_row_joined:
                return True

            # Fallback heuristic: look for overtime-related headers and sane layout
            indicators_found = 0
            header_keywords = ['overtime', 'overtime rate', 'rate', 'employee', 'employee name']
            header_found = False
            for row in range(1, min(6, worksheet.max_row + 1)):
                for col in range(1, min(8, worksheet.max_column + 1)):
                    cell_value = self._cell_value_to_string(worksheet.cell(row=row, column=col).value).lower()
                    for kw in header_keywords:
                        if kw in cell_value:
                            header_found = True
                            break
                    if header_found:
                        break
                if header_found:
                    indicators_found += 1
                    break

            # Check for employee names in configured columns starting from configured row
            valid_employees = 0
            for row in range(OVERTIME_RATES_CONFIG["data_start_row"], min(12, worksheet.max_row + 1)):
                employee_name = None
                for col in OVERTIME_RATES_CONFIG["employee_name_columns"]:
                    employee_name = self._cell_value_to_string(worksheet.cell(row=row, column=col).value)
                    if employee_name:
                        break
                if employee_name and len(employee_name.split()) >= 2:
                    valid_employees += 1
            if valid_employees >= 1:
                indicators_found += 1

            # Check for Yes/No pattern in nearby columns
            yes_no_found = 0
            for row in range(2, min(12, worksheet.max_row + 1)):
                for col in range(2, min(6, worksheet.max_column + 1)):
                    flag_value = self._cell_value_to_string(worksheet.cell(row=row, column=col).value).lower()
                    if flag_value in ['yes', 'no', 'y', 'n', 'true', 'false', '1', '0']:
                        yes_no_found += 1
                        break
            if yes_no_found >= 1:
                indicators_found += 1

            return indicators_found >= 2
        except Exception:
            return False

    def parse(self, file_path: str) -> Dict[str, Any]:
        """
        Parse the overtime rates file and return structured data.
        
        Args:
            file_path: Path to the overtime rates file
            
        Returns:
            Dictionary containing parsed overtime rates data in standard format
            
        Raises:
            FileNotFoundError: If the file doesn't exist
            ValueError: If the file format is invalid or data is malformed
        """
        self._validate_file_exists(file_path)
        
        try:
            # 1. Initialize empty containers for aggregated results
            master_employees_data = {}
            total_entries = 0
            # Track where the first occurrence for each employee was kept
            kept_occurrence = {}

            # 2. Load the workbook from file_path
            workbook = self._load_excel_workbook(file_path)

            # First, try to find sheets whose first row contains the exact title phrase
            candidate_sheets = []
            for worksheet in workbook.worksheets:
                first_row = ' '.join(
                    [self._cell_value_to_string(worksheet.cell(row=1, column=c).value).lower()
                     for c in range(1, min(worksheet.max_column, OVERTIME_RATES_CONFIG["max_search_columns"]) + 1)]
                )
                if OVERTIME_RATES_CONFIG["exact_title_phrase"] in first_row:
                    candidate_sheets.append(worksheet)

            # If none found by exact phrase, fall back to heuristic detection
            if not candidate_sheets:
                for worksheet in workbook.worksheets:
                    if self._is_valid_overtime_sheet(worksheet):
                        candidate_sheets.append(worksheet)

            # For overtime rates, do NOT merge multiple sheets. Choose the first matching sheet only.
            chosen_sheets = []
            if candidate_sheets:
                chosen_sheets = [candidate_sheets[0]]

            # Parse chosen sheets with diagnostics
            sheet_errors = {}
            sheets_parsed_successfully = 0
            parsed_sheet_names = []
            skipped_sheets = []
            duplicates = []

            for worksheet in chosen_sheets:
                try:
                    for row_num in range(OVERTIME_RATES_CONFIG["data_start_row"], worksheet.max_row + 1):
                        employee_name = None
                        for col in OVERTIME_RATES_CONFIG["employee_name_columns"]:
                            employee_name = self._cell_value_to_string(worksheet.cell(row=row_num, column=col).value)
                            if employee_name:
                                break
                        if not employee_name:
                            continue

                        if employee_name in master_employees_data:
                            # record duplicate and keep first occurrence
                            kept = kept_occurrence.get(employee_name, {})
                            duplicates.append({
                                'employee_name': employee_name,
                                'kept_sheet': kept.get('sheet'),
                                'kept_row': kept.get('row'),
                                'skipped_sheet': worksheet.title,
                                'skipped_row': row_num
                            })
                            continue

                        different_rate_flag = None
                        overtime_rate_value = None
                        
                        # Try configured columns for overtime flag and rate
                        for i, flag_col in enumerate(OVERTIME_RATES_CONFIG["overtime_flag_columns"]):
                            flag_value = self._cell_value_to_string(worksheet.cell(row=row_num, column=flag_col).value).lower()
                            if flag_value:
                                different_rate_flag = flag_value
                                # Use corresponding rate column
                                if i < len(OVERTIME_RATES_CONFIG["overtime_rate_columns"]):
                                    overtime_rate_value = worksheet.cell(row=row_num, column=OVERTIME_RATES_CONFIG["overtime_rate_columns"][i]).value
                                break

                        has_different_rate = different_rate_flag in ['yes', 'y', 'true', '1']

                        overtime_rate = None
                        if has_different_rate:
                            rate_str = self._cell_value_to_string(overtime_rate_value)
                            if rate_str:
                                rate_str = rate_str.replace(',', '.')
                                try:
                                    rate_val = float(rate_str)
                                    if rate_val > 0:
                                        overtime_rate = rate_val
                                except (ValueError, TypeError):
                                    has_different_rate = False

                        master_employees_data[employee_name] = {
                            "employee_name": employee_name,
                            "has_different_overtime_rate": has_different_rate,
                            "overtime_rate": overtime_rate
                        }
                        # record first-kept occurrence
                        kept_occurrence[employee_name] = {'sheet': worksheet.title, 'row': row_num}
                        total_entries += 1
                    sheets_parsed_successfully += 1
                    parsed_sheet_names.append(worksheet.title)
                except Exception as e:
                    sheet_errors[worksheet.title] = str(e)
                    continue

            # Determine skipped sheets for diagnostics
            chosen_title_set = set(parsed_sheet_names)
            for ws in workbook.worksheets:
                if ws.title not in chosen_title_set:
                    skipped_sheets.append(ws.title)

            workbook.close()

            # Return empty structure if nothing parsed
            if not master_employees_data:
                return {
                    "file_type": "overtime_rates",
                    "pay_period_end_date": date.today(),
                    "employees": [],
                    "metadata": {
                        "source_file": file_path,
                        "parsed_at": datetime.now(),
                        "total_employees": 0,
                        "total_entries": 0
                    },
                    "overtime_rates_lookup": {}
                }

            employees_list = [{"employee_name": name, "entries": []} for name in master_employees_data.keys()]

            # Attach diagnostics into metadata so orchestrator can report them
            metadata = {
                "source_file": file_path,
                "parsed_at": datetime.now(),
                "total_employees": len(master_employees_data),
                "total_entries": total_entries,
                "diagnostics": {
                    "parsed_sheets": parsed_sheet_names,
                    "skipped_sheets": skipped_sheets,
                    "sheet_errors": sheet_errors,
                    "duplicates": duplicates
                }
            }

            return {
                "file_type": "overtime_rates",
                "pay_period_end_date": date.today(),
                "employees": employees_list,
                "metadata": metadata,
                "overtime_rates_lookup": master_employees_data
            }
        except Exception as e:
            raise ValueError(f"Failed to parse overtime rates file {file_path}: {str(e)}")